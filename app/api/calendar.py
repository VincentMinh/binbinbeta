from fastapi import APIRouter, Request, Depends
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
import os
from collections import defaultdict, OrderedDict
from typing import Optional
import calendar
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from datetime import datetime, date, timedelta

from ..db.session import get_db
from ..db.models import User, AttendanceRecord, ServiceRecord, Branch, Department
from ..core.security import get_active_branch
from ..core.utils import VN_TZ
from ..core.config import ROLE_MAP
from sqlalchemy import cast, Date
from sqlalchemy.orm import joinedload

from fastapi.templating import Jinja2Templates

router = APIRouter()

# Xác định đường dẫn tuyệt đối đến thư mục gốc của project 'app'
APP_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

# Tạo đường dẫn tuyệt đối đến thư mục templates
templates = Jinja2Templates(directory=os.path.join(APP_ROOT, "templates"))


@router.get("/calendar-view", response_class=HTMLResponse)
def view_attendance_calendar(
    request: Request,
    db: Session = Depends(get_db),
    chi_nhanh: Optional[str] = None, # Đây là branch_code
    month: Optional[int] = None,
    year: Optional[int] = None,
):
    user_data = request.session.get("user")
    if not user_data or user_data.get("role") not in ["admin", "boss", "quanly", "letan", "ktv"]:
        return RedirectResponse("/choose-function", status_code=303)

    # Lấy danh sách chi nhánh và phòng ban từ DB để hiển thị trong bộ lọc
    # Sửa lỗi: Loại bỏ Admin và Boss khỏi danh sách chi nhánh
    all_branches_obj = db.query(Branch).filter(Branch.branch_code.notin_(['Admin', 'Boss'])).all()

    # Logic sắp xếp chi nhánh tùy chỉnh
    b_branches = []
    other_branches = []
    for b in all_branches_obj:
        if b.branch_code.startswith('B') and b.branch_code[1:].isdigit():
            b_branches.append(b.branch_code)
        else:
            other_branches.append(b.branch_code)

    b_branches.sort(key=lambda x: int(x[1:]))
    other_branches.sort()
    display_branches = b_branches + other_branches

    # Xử lý giá trị chi_nhanh mặc định. Nếu chi_nhanh là chuỗi rỗng từ form, nó sẽ trở thành None.
    # Nếu không có chi nhánh nào được chọn (kể cả lần đầu vào trang), thì xử lý mặc định.
    if chi_nhanh is None:
        user_role = user_data.get("role")
        if user_role in ["ktv", "quanly"]:
            chi_nhanh = user_data.get("branch", "B1")
        elif user_role == "letan":
            active_branch = get_active_branch(request, db, user_data)
            chi_nhanh = active_branch or user_data.get("branch", "B1")
        elif user_role in ["admin", "boss"]:
            # Để trống để mặc định là "Tất cả chi nhánh"
            chi_nhanh = "B1"

    now = datetime.now(VN_TZ)
    current_month = month if month else now.month
    current_year = year if year else now.year

    start_date_of_month = date(current_year, current_month, 1)
    end_date_of_month = date(current_year, current_month, calendar.monthrange(current_year, current_month)[1])
    
    _, num_days = calendar.monthrange(current_year, current_month)
    
    employee_data = defaultdict(lambda: {
        "name": "", "role": "", "role_key": "", "main_branch": "",
        "worked_away_from_main_branch": False,
        "daily_work": defaultdict(lambda: {"work_units": 0, "is_overtime": False, "work_branch": "", "services": []})
    })

    if chi_nhanh:
        # === BƯỚC 1: LẤY DANH SÁCH NHÂN VIÊN CHÍNH THỨC CỦA VIEW HIỆN TẠI ===
        base_employee_query = db.query(User).options(joinedload(User.department), joinedload(User.main_branch))
        
        role_map_filter = {"KTV": "ktv", "Quản lý": "quanly"}
        code_prefix_filter = {"LTTC": "LTTC", "BPTC": "BPTC"}

        if chi_nhanh in role_map_filter:
            base_employee_query = base_employee_query.join(User.department).filter(Department.role_code == role_map_filter[chi_nhanh])
        elif chi_nhanh in code_prefix_filter:
            base_employee_query = base_employee_query.filter(User.employee_code.startswith(code_prefix_filter[chi_nhanh]))
        else: # Lọc theo chi nhánh thông thường
            # Sửa: Nếu chi_nhanh là chuỗi rỗng (Tất cả), không lọc theo chi nhánh chính
            if chi_nhanh:
                base_employee_query = base_employee_query.join(User.main_branch).filter(Branch.branch_code == chi_nhanh)
        
        
        base_employees = base_employee_query.all()
        for emp in base_employees:
            emp_code = emp.employee_code
            if emp_code not in employee_data:
                employee_data[emp_code]["name"] = emp.name
                employee_data[emp_code]["main_branch"] = emp.main_branch.branch_code if emp.main_branch else ''
                role_code = emp.department.role_code if emp.department else 'khac'
                employee_data[emp_code]["role_key"] = role_code # Giữ role_code để sắp xếp
                employee_data[emp_code]["role"] = ROLE_MAP.get(role_code, role_code)

        # === BƯỚC 2: LẤY TẤT CẢ BẢN GHI CHẤM CÔNG VÀ DỊCH VỤ LIÊN QUAN ===
        # Query cho điểm danh
        att_q = db.query(AttendanceRecord).options(joinedload(AttendanceRecord.branch))
        # Query cho dịch vụ
        svc_q = db.query(ServiceRecord).options(joinedload(ServiceRecord.branch))
        
        # Áp dụng bộ lọc
        # Áp dụng bộ lọc
        if chi_nhanh in role_map_filter:
            # Lấy role_code (ví dụ: "ktv") từ bộ lọc
            role_code_to_filter = role_map_filter[chi_nhanh]
            
            # Lấy role_name (ví dụ: "Kỹ thuật viên") từ ROLE_MAP
            # Dùng .get() để tránh lỗi nếu không tìm thấy, dù trường hợp này không nên xảy ra
            role_name_to_filter = ROLE_MAP.get(role_code_to_filter, role_code_to_filter)

            # Lọc theo TÊN CHỨC VỤ (đã sửa) thay vì MÃ CHỨC VỤ
            att_q = att_q.filter(AttendanceRecord.role_snapshot == role_name_to_filter)
            svc_q = svc_q.filter(ServiceRecord.role_snapshot == role_name_to_filter)
        elif chi_nhanh in code_prefix_filter:
            att_q = att_q.filter(AttendanceRecord.employee_code_snapshot.startswith(code_prefix_filter[chi_nhanh]))
            svc_q = svc_q.filter(ServiceRecord.employee_code_snapshot.startswith(code_prefix_filter[chi_nhanh]))
        else:
            branch_to_filter_obj = next((b for b in all_branches_obj if b.branch_code == chi_nhanh), None)
            # Sửa: Chỉ lọc theo chi nhánh nếu có chi nhánh được chọn
            if chi_nhanh and branch_to_filter_obj:
                att_q = att_q.filter(AttendanceRecord.branch_id == branch_to_filter_obj.id)
                svc_q = svc_q.filter(ServiceRecord.branch_id == branch_to_filter_obj.id)

        # Lọc theo khoảng thời gian
        att_records = att_q.filter(cast(AttendanceRecord.attendance_datetime, Date).between(start_date_of_month, end_date_of_month)).all()
        svc_records = svc_q.filter(cast(ServiceRecord.service_datetime, Date).between(start_date_of_month, end_date_of_month)).all()

        all_records = att_records + svc_records

        # === BƯỚC 3: XỬ LÝ DỮ LIỆU ĐỂ HIỂN THỊ ===
        for rec in all_records:
            is_att = isinstance(rec, AttendanceRecord)
            dt = rec.attendance_datetime if is_att else rec.service_datetime
            
            # Chuyển đổi thời gian DB (UTC) về giờ Việt Nam (GMT+7)
            dt_local = dt.astimezone(VN_TZ)
            # Áp dụng logic trên giờ Việt Nam
            work_date = dt_local.date() - timedelta(days=1) if dt_local.hour < 7 else dt_local.date()

            if work_date.month != current_month or work_date.year != current_year:
                continue

            day_of_month = work_date.day
            emp_code = rec.employee_code_snapshot

            # Nếu nhân viên chưa có trong danh sách (trường hợp tăng ca từ chi nhánh khác)
            if emp_code not in employee_data:
                employee_data[emp_code]["name"] = rec.employee_name_snapshot
                employee_data[emp_code]["main_branch"] = rec.main_branch_snapshot
                employee_data[emp_code]["role_key"] = rec.role_snapshot
                employee_data[emp_code]["role"] = ROLE_MAP.get(rec.role_snapshot, rec.role_snapshot)

            # Sửa lỗi logic: so sánh branch_code của nơi làm việc với branch_code của chi nhánh chính đã lưu
            main_branch_of_employee = employee_data[emp_code].get("main_branch")
            if rec.branch and main_branch_of_employee and rec.branch.branch_code != main_branch_of_employee:
                employee_data[emp_code]["worked_away_from_main_branch"] = True

            daily_work_entry = employee_data[emp_code]["daily_work"][day_of_month]
            if is_att:
                daily_work_entry["work_units"] += rec.work_units or 0
            else: # Service Record
                service_summary = daily_work_entry.setdefault("service_summary", defaultdict(int))
                service_summary[rec.service_type] += rec.quantity or 0

        # Chuyển đổi service_summary thành list string để dễ render
        for emp_code in employee_data:
            for day_data in employee_data[emp_code]["daily_work"].values():
                if "service_summary" in day_data:
                    summary = day_data.pop("service_summary")
                    day_data["services"] = [f"{k}: {v}" for k, v in summary.items()]

        # === BƯỚC 4: TỐI ƯU HÓA - LẤY DỮ LIỆU THỐNG KÊ MỘT LẦN ===
        # 1. Xác định danh sách nhân viên chính của view này để lấy dữ liệu
        main_employee_codes = [
            emp_code for emp_code, emp_details in employee_data.items()
            if (
                emp_details.get("main_branch") == chi_nhanh
                or (chi_nhanh == "KTV" and emp_details.get("role_key") == "ktv")
                or (chi_nhanh == "Quản lý" and emp_details.get("role_key") == "quanly")
                or (chi_nhanh == "LTTC" and emp_details.get("role_key") == "lttc")
                or (chi_nhanh == "BPTC" and emp_details.get("role_key") == "bptc")
            )
        ]

        all_atts_for_stats = defaultdict(list)
        all_services_for_stats = defaultdict(list)

        if main_employee_codes:
            # 2. Lấy tất cả bản ghi điểm danh và dịch vụ cho các nhân viên đó trong khoảng thời gian liên quan
            start_query_date = date(current_year, current_month, 1)
            end_query_date = date(current_year, current_month, num_days) + timedelta(days=1)

            all_atts_raw_bulk = db.query(AttendanceRecord).options(joinedload(AttendanceRecord.branch)).filter(
                AttendanceRecord.employee_code_snapshot.in_(main_employee_codes),
                cast(AttendanceRecord.attendance_datetime, Date).between(start_query_date, end_query_date)
            ).all()

            all_services_bulk = db.query(ServiceRecord).options(joinedload(ServiceRecord.branch)).filter(
                ServiceRecord.employee_code_snapshot.in_(main_employee_codes),
                cast(ServiceRecord.service_datetime, Date).between(start_date_of_month, end_date_of_month)
            ).all()

            # 3. Nhóm các bản ghi theo mã nhân viên để tra cứu nhanh
            for att in all_atts_raw_bulk:
                all_atts_for_stats[att.employee_code_snapshot].append(att)
            for svc in all_services_bulk:
                all_services_for_stats[svc.employee_code_snapshot].append(svc)

        # === BƯỚC 4: TÍNH TOÁN THỐNG KÊ CHO DASHBOARD (ĐÃ BỊ THIẾU) ===
        for emp_code, emp_details in employee_data.items():
            # Chỉ tính cho nhân viên có chi nhánh chính là chi nhánh đang xem
            is_main_employee_of_view = (
                emp_details.get("main_branch") == chi_nhanh
                or (chi_nhanh == "KTV" and emp_details.get("role_key") == "ktv")
                or (chi_nhanh == "Quản lý" and emp_details.get("role_key") == "quanly")
                or (chi_nhanh == "LTTC" and emp_details.get("role_key") == "lttc")
                or (chi_nhanh == "BPTC" and emp_details.get("role_key") == "bptc")
            )

            if is_main_employee_of_view:
                # --- TÍNH TOÁN DASHBOARD ---
                # 1. Lấy dữ liệu đã được truy vấn sẵn, không query lại DB
                all_atts_raw = all_atts_for_stats.get(emp_code, [])

                # Helper để xác định ngày làm việc (ca đêm < 7h sáng tính cho ngày hôm trước)
                def get_work_day(att_datetime):
                    # Chuyển đổi thời gian DB (UTC) về giờ Việt Nam (GMT+7)
                    dt_local = att_datetime.astimezone(VN_TZ)
                    # Áp dụng logic trên giờ Việt Nam
                    return dt_local.date() - timedelta(days=1) if dt_local.hour < 7 else dt_local.date()
                # Gắn "work_day" vào mỗi bản ghi và lọc lại theo tháng đang xem
                all_atts = [
                    {
                        "work_day": get_work_day(att.attendance_datetime),
                        **att.__dict__
                    }
                    for att in all_atts_raw
                ]
                all_atts = [
                    att for att in all_atts 
                    if att["work_day"].month == current_month and att["work_day"].year == current_year
                ]

                # 2. Xử lý dữ liệu điểm danh dựa trên "work_day"
                tong_so_cong = 0.0
                work_days_set = set()
                overtime_work_days_set = set()
                daily_work_units = defaultdict(float)

                for att in all_atts:
                    work_day = att['work_day']
                    so_cong = att.get('work_units') or 0
                    tong_so_cong += so_cong
                    if so_cong > 0:
                        work_days_set.add(work_day)
                    daily_work_units[work_day] += so_cong
                    if att.get('is_overtime'):
                        overtime_work_days_set.add(work_day)

                # Xác định ngày tăng ca dựa trên tổng công > 1
                for day, total_units in daily_work_units.items():
                    if total_units > 1:
                        overtime_work_days_set.add(day)

                # Lấy chi tiết tăng ca
                overtime_details = []
                main_branch = emp_details.get("main_branch")

                # Xác định những ngày làm việc có chấm công ở chi nhánh khác (với số công > 0)
                other_branch_work_days = {
                    # Sửa lỗi: Tính work_day trực tiếp từ att.attendance_datetime
                    get_work_day(att.attendance_datetime)
                    for att in all_atts_raw
                    if main_branch and att.branch and att.branch.branch_code != main_branch and (att.work_units or 0) > 0
                }

                # Set để đảm bảo mỗi ngày chỉ xử lý 1 lần cho trường hợp >1 công
                processed_main_branch_overtime_days = set()

                # Lặp qua tất cả các bản ghi để xây dựng chi tiết
                for att in all_atts:
                    work_day = att.get('work_day')

                    # Bỏ qua nếu không phải là ngày tăng ca
                    if work_day not in overtime_work_days_set:
                        continue

                    # Ưu tiên 1: Tăng ca do đi chi nhánh khác
                    # Sửa lỗi: att.get('branch') không tồn tại, phải dùng att['branch_id']
                    if work_day in other_branch_work_days:
                        # Chỉ thêm các bản ghi ở chi nhánh khác (có công)
                        if main_branch and att.get('_sa_instance_state').object.branch.branch_code != main_branch and (att.get('work_units') or 0) > 0:
                            # Sửa lỗi: Chuyển đổi sang múi giờ Việt Nam trước khi định dạng
                            local_time = att['attendance_datetime'].astimezone(VN_TZ)
                            overtime_details.append({ 
                                "date": local_time.strftime('%d/%m/%Y'), "time": local_time.strftime('%H:%M'), 
                                "branch": att.get('_sa_instance_state').object.branch.branch_code, "work_units": att.get('work_units') })
                    # Trường hợp 2: Tăng ca do làm >1 công (và chỉ làm tại chi nhánh chính)
                    elif daily_work_units.get(work_day, 0) > 1:
                        if work_day not in processed_main_branch_overtime_days:
                            # Chỉ hiển thị 1 dòng tóm tắt cho ngày này
                            overtime_details.append({ "date": work_day.strftime('%d/%m/%Y'), "time": "Nhiều ca", "branch": main_branch, "work_units": f"{daily_work_units.get(work_day, 0):.1f}" })
                            processed_main_branch_overtime_days.add(work_day)

                # 3. Lấy dữ liệu dịch vụ đã được truy vấn sẵn
                all_services = all_services_for_stats.get(emp_code, [])

                # 4. Tổng hợp kết quả
                so_ngay_lam = len(work_days_set)
                so_ngay_tang_ca = len(overtime_work_days_set)

                # --- LOGIC MỚI CHO SỐ NGÀY NGHỈ ---
                is_current_month_view = (current_year == now.year and current_month == now.month)
                
                if is_current_month_view:
                    # Đối với tháng hiện tại, số ngày nghỉ được tính từ đầu tháng đến ngày hôm nay.
                    days_passed = now.day
                    # Lọc ra những ngày đã làm việc tính đến hôm nay.
                    worked_days_so_far = {d for d in work_days_set if d <= now.date()}
                    so_ngay_nghi = days_passed - len(worked_days_so_far)
                else:
                    # Đối với các tháng trong quá khứ, tính như cũ.
                    so_ngay_nghi = num_days - so_ngay_lam
                so_ngay_nghi = max(0, so_ngay_nghi)
                laundry_details = []
                ironing_details = []
                tong_dich_vu_giat = 0
                tong_dich_vu_ui = 0

                for svc in all_services:
                    try:
                        quantity = int(svc.quantity)
                    except (ValueError, TypeError):
                        quantity = 0
                    
                    # Sửa lỗi: Chuyển đổi sang múi giờ Việt Nam trước khi định dạng
                    local_time = svc.service_datetime.astimezone(VN_TZ)
                    detail = {
                        "date": local_time.strftime('%d/%m/%Y'), 
                        "time": local_time.strftime('%H:%M'),
                        "branch": svc.branch.branch_code if svc.branch else '', "room": svc.room_number, "quantity": svc.quantity
                    }

                    if svc.service_type == 'Giặt':
                        tong_dich_vu_giat += quantity
                        laundry_details.append(detail)
                    elif svc.service_type == 'Ủi':
                        tong_dich_vu_ui += quantity
                        ironing_details.append(detail)

                emp_details["dashboard_stats"] = {
                    "so_ngay_lam": so_ngay_lam,
                    "so_ngay_nghi": so_ngay_nghi,
                    "so_ngay_tang_ca": so_ngay_tang_ca,
                    "tong_so_cong": tong_so_cong,
                    "tong_dich_vu_giat": tong_dich_vu_giat,
                    "tong_dich_vu_ui": tong_dich_vu_ui,
                    "overtime_details": sorted(overtime_details, key=lambda x: datetime.strptime(x['date'], '%d/%m/%Y')),
                    "laundry_details": sorted(laundry_details, key=lambda x: datetime.strptime(x['date'], '%d/%m/%Y')),
                    "ironing_details": sorted(ironing_details, key=lambda x: datetime.strptime(x['date'], '%d/%m/%Y')),
                }

    # Sắp xếp nhân viên
    role_priority = {"letan": 0, "buongphong": 1, "baove": 2, "ktv": 3, "quanly": 4}
    sorted_employee_list = sorted(
        employee_data.items(),
        key=lambda item: (
            role_priority.get(item[1].get("role_key", "khac"), 99),
            item[1].get("name", "")
        )
    )
    sorted_employee_data = OrderedDict(sorted_employee_list)

    return templates.TemplateResponse("calendar_view.html", {
        "request": request,
        "user": user_data,
        "branches": display_branches,
        "selected_branch": chi_nhanh,
        "selected_month": current_month,
        "selected_year": current_year,
        "num_days": num_days,
        "employee_data": sorted_employee_data,
        "employee_data_for_js": sorted_employee_data, # Sửa lỗi: Cung cấp dữ liệu cho dashboard
        "current_day": now.day if now.month == current_month and now.year == current_year else None,
    })

@router.get("/calendar-export-excel")
def export_attendance_calendar_excel(
    request: Request,
    db: Session = Depends(get_db),
    month: Optional[int] = None,
    year: Optional[int] = None,
):
    """
    Endpoint chuyên dụng để xuất file Excel chấm công theo dạng lịch
    với 2 hàng (công chính, tăng ca) cho mỗi nhân viên.
    """
    user_data = request.session.get("user")
    if not user_data or user_data.get("role") not in ["admin", "boss", "quanly"]:
        return RedirectResponse("/choose-function", status_code=303)

    now = datetime.now(VN_TZ)
    current_month = month if month else now.month
    current_year = year if year else now.year

    start_date_of_month = date(current_year, current_month, 1)
    _, num_days = calendar.monthrange(current_year, current_month)
    # Sửa lỗi logic: cần lấy đến 7h sáng ngày 1 tháng sau
    end_date_of_month_query = date(current_year, current_month, num_days) + timedelta(days=1)

    # === 1. LẤY VÀ SẮP XẾP TẤT CẢ NHÂN VIÊN ===
    # Yêu cầu là xuất tất cả, sắp xếp B1 -> B2..., nên chúng ta bỏ qua bộ lọc 'chi_nhanh'

    all_users = db.query(User).options(
        joinedload(User.department), 
        joinedload(User.main_branch)
    ).filter(
        User.employee_code.notin_(['admin', 'boss']) # Loại trừ user hệ thống
    ).all()

    # Hàm sort key phức tạp để đảm bảo B1 -> B2 -> B10 -> Khác
    def get_sort_key(user):
        branch_code = user.main_branch.branch_code if user.main_branch else 'ZZZ'
        role_code = user.department.role_code if user.department else 'z'

        # Ưu tiên 1: Sắp xếp chi nhánh (B1, B2, ..., B10, ..., Khác)
        if branch_code.startswith('B') and branch_code[1:].isdigit():
            branch_sort_key = (0, int(branch_code[1:]))
        else:
            branch_sort_key = (1, branch_code)

        # Ưu tiên 2: Sắp xếp theo vai trò (như trong code gốc)
        role_priority = {"letan": 0, "buongphong": 1, "baove": 2, "ktv": 3, "quanly": 4}
        role_sort_key = role_priority.get(role_code, 99)

        return (branch_sort_key, role_sort_key, user.name)

    all_users.sort(key=get_sort_key)

    # Tạo map tra cứu chi nhánh chính của user
    user_main_branch_map = {u.employee_code: u.main_branch.branch_code if u.main_branch else '' for u in all_users}

    # === 2. LẤY TẤT CẢ DỮ LIỆU CHẤM CÔNG TRONG THÁNG ===

    all_att_records = db.query(AttendanceRecord).options(joinedload(AttendanceRecord.branch)).filter(
        AttendanceRecord.attendance_datetime >= start_date_of_month,
        AttendanceRecord.attendance_datetime < end_date_of_month_query
    ).all()

    # === 3. XỬ LÝ DỮ LIỆU (PIVOT) ===
    # Cấu trúc: data_pivot[emp_code][day_num]["main_work" | "overtime_work"]
    data_pivot = defaultdict(lambda: 
        defaultdict(lambda: {
            "main_work": 0.0,
            "overtime_work": 0.0
        })
    )

    for rec in all_att_records:
        # Logic ca đêm (lấy từ code gốc của bạn)
        dt_local = rec.attendance_datetime.astimezone(VN_TZ)
        work_date = dt_local.date() - timedelta(days=1) if dt_local.hour < 7 else dt_local.date()

        # Chỉ xử lý các ngày trong tháng được chọn
        if work_date.month != current_month or work_date.year != current_year:
            continue

        day_num = work_date.day
        emp_code = rec.employee_code_snapshot

        main_branch = user_main_branch_map.get(emp_code)
        work_branch = rec.branch.branch_code if rec.branch else ''

        work_units = rec.work_units or 0
        is_ot_branch = (main_branch and work_branch and work_branch != main_branch)

        day_entry = data_pivot[emp_code][day_num]

        # Phân loại công chính và tăng ca
        if rec.is_overtime or is_ot_branch:
            day_entry["overtime_work"] += work_units
        else:
            day_entry["main_work"] += work_units

    # === 4. TẠO FILE EXCEL VỚI OPENPYXL ===

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Chấm công T{current_month}-{current_year}"

    # --- Định nghĩa Styles ---
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fill màu giống hình ảnh
    main_work_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid") # Xanh lá nhạt
    ot_work_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Xám nhạt

    cn_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Đỏ đậm cho Chủ Nhật
    cn_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # --- Tạo hàng Header ---
    headers = ["STT", "TÊN NHÂN VIÊN"] + [f"{d:02d}" for d in range(1, num_days + 1)] + ["TỔNG CỘNG"]
    ws.append(headers)

    # Style hàng Header
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        # Highlight Chủ Nhật (CN)
        if col_idx > 2 and col_idx <= num_days + 2:
            day_num = col_idx - 2
            day_date = date(current_year, current_month, day_num)
            if day_date.weekday() == 6: # 6 = Sunday
                cell.fill = cn_fill
                cell.font = cn_font
                cell.value = f"CN\n{day_num:02d}" # Giống hình ảnh

    # --- Ghi dữ liệu nhân viên ---
    current_row = 2
    stt = 1
    for user in all_users:
        emp_code = user.employee_code
        emp_name_with_branch = f"{user.name}_{user.main_branch.branch_code}" if user.main_branch else user.name

        main_row_data = [stt, emp_name_with_branch]
        ot_row_data = ["", "Tăng ca"]

        total_main_work = 0.0
        total_ot_work = 0.0

        for d in range(1, num_days + 1):
            day_data = data_pivot[emp_code].get(d)
            main_work = 0.0
            ot_work = 0.0

            if day_data:
                main_work = day_data["main_work"]
                ot_work = day_data["overtime_work"]

            # Hiển thị số nếu > 0, ngược lại để trống
            main_row_data.append(main_work if main_work > 0 else "")
            ot_row_data.append(ot_work if ot_work > 0 else "")

            total_main_work += main_work
            total_ot_work += ot_work

        # Thêm cột tổng cộng
        main_row_data.append(total_main_work if total_main_work > 0 else "")
        ot_row_data.append(total_ot_work if total_ot_work > 0 else "")

        # Ghi vào sheet
        ws.append(main_row_data)
        ws.append(ot_row_data)

        # --- Style cho 2 hàng vừa thêm ---
        main_data_row = ws[current_row]
        ot_data_row = ws[current_row + 1]

        for col_idx in range(1, len(headers) + 1):
            main_cell = main_data_row[col_idx - 1]
            ot_cell = ot_data_row[col_idx - 1]

            main_cell.border = thin_border
            ot_cell.border = thin_border
            main_cell.alignment = center_align
            ot_cell.alignment = center_align

            if col_idx > 2 and col_idx <= num_days + 2:
                main_cell.fill = main_work_fill
                ot_cell.fill = ot_work_fill

        # Merge ô STT
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        # Merge ô Tên
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2)

        # Căn chỉnh lại ô STT và Tên sau khi merge
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=current_row + 1, column=2).alignment = Alignment(horizontal="right", vertical="center")


        current_row += 2
        stt += 1

    # --- Điều chỉnh độ rộng cột ---
    ws.column_dimensions['A'].width = 5  # STT
    ws.column_dimensions['B'].width = 30 # TÊN
    for d in range(1, num_days + 1):
        ws.column_dimensions[get_column_letter(d + 2)].width = 5 # Các cột ngày
    ws.column_dimensions[get_column_letter(num_days + 3)].width = 12 # TỔNG CỘNG

    # Đóng băng (Freeze) hàng header và cột tên
    ws.freeze_panes = "C2"

    # --- Lưu vào stream và trả về ---
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    filename = f"ChamCong_Thang_{current_month}_{current_year}.xlsx"

    return StreamingResponse(
        output_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )