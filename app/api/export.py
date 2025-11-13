from fastapi import APIRouter, Request, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from urllib.parse import quote

from ..db.session import get_db
from ..db.models import Task
from ..core.utils import VN_TZ, format_datetime_display
from ..core.config import logger

# Import các hàm query đã được module hóa
from .tasks import _get_filtered_tasks_query
from .results import _get_filtered_records_query

router = APIRouter()

def _auto_adjust_worksheet_columns(worksheet):
    """Helper function to adjust column widths of a worksheet."""
    for i, column_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        # Also check header length
        header_cell = worksheet.cell(row=1, column=i)
        if header_cell.value:
            max_length = len(str(header_cell.value))

        for cell in column_cells:
            if cell.row == 1: continue # Skip header
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

@router.get("/api/tasks/export-excel", tags=["Export"])
async def export_tasks_to_excel(
    request: Request,
    chi_nhanh: str = "",
    search: str = "",
    trang_thai: str = "",
    han_hoan_thanh: str = "",
    db: Session = Depends(get_db)
):
    user_data = request.session.get("user")
    if not user_data or user_data.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")

    tasks_query = _get_filtered_tasks_query(db, user_data, chi_nhanh, search, trang_thai, han_hoan_thanh)
    rows_all = tasks_query.order_by(Task.due_date.nullslast()).all()
    if not rows_all:
        return Response(status_code=204, content="Không có dữ liệu để xuất.")

    data_for_export = [{
        "ID": t.id,
        "Chi Nhánh": t.branch.name if t.branch else '',
        "Phòng": t.room_number,
        "Mô Tả": t.description,
        "Ngày Tạo": format_datetime_display(t.created_at, with_time=True),
        "Hạn Hoàn Thành": format_datetime_display(t.due_date, with_time=False),
        "Trạng Thái": t.status,
        "Người Tạo": t.author.name if t.author else '',
        "Người Thực Hiện": t.assignee.name if t.assignee else '',
        "Ngày Hoàn Thành": format_datetime_display(t.completed_at, with_time=True) if t.completed_at else "",
        "Ghi Chú": t.notes or "",
    } for t in rows_all]

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CongViec"

    headers = list(data_for_export[0].keys())
    ws.append(headers)
    for row_data in data_for_export:
        ws.append(list(row_data.values()))
    _auto_adjust_worksheet_columns(ws)

    wb.save(output)
    output.seek(0)

    # Lấy kích thước của file trong memory để thêm vào header Content-Length
    file_size = output.getbuffer().nbytes

    filename = f"danh_sach_cong_viec_{datetime.now(VN_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    # Mã hóa tên file để tương thích với nhiều trình duyệt hơn
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Length": str(file_size)
        }
    )

@router.get("/api/attendance/export-excel", tags=["Export"])
async def export_attendance_to_excel(request: Request, db: Session = Depends(get_db)):
    user_data = request.session.get("user")
    if not user_data or user_data.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")

    # Sử dụng hàm query đã được module hóa từ results.py
    query, columns = _get_filtered_records_query(db, request.query_params, user_data)
    records = db.execute(query).all()

    if not records:
        return Response(status_code=204, content="Không có dữ liệu để xuất.")

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DiemDanh"

    # Ghi header
    ws.append([col.name for col in columns])
    # Ghi dữ liệu
    for rec in records:
        ws.append([format_datetime_display(val) if isinstance(val, datetime) else val for val in rec])
    
    _auto_adjust_worksheet_columns(ws)
    wb.save(output)
    output.seek(0)
    
    # Lấy kích thước của file trong memory để thêm vào header Content-Length
    file_size = output.getbuffer().nbytes
    
    filename = f"ket_qua_diem_danh_{datetime.now(VN_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    # Mã hóa tên file để tương thích với nhiều trình duyệt hơn
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Length": str(file_size)
        }
    )